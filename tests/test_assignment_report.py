"""Tests for the assignment validation report and the Excel generator."""
import openpyxl

from application.services.credit_system import (
    AssignmentReporter,
    CreditSystem,
)
from domain.entities.group import Group, LabSession
from domain.entities.professor import Professor
from domain.value_objects import TimeSlot, WeekWindow
from infrastructure.excel.excel_generator import ExcelGenerator


def _slot(day_idx=2):
    days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    return TimeSlot(day_idx=day_idx, day=days[day_idx], block_id=1,
                    block_label="08:30-10:30")


# --- AssignmentReporter ------------------------------------------------------
def test_reporter_stats_balanced_and_overload():
    profs = [
        Professor(name="A", practice_credits_by_subject={"S1": 6}),  # expected 30
        Professor(name="B", practice_credits_by_subject={"S2": 2}),  # expected 10
    ]
    cs = CreditSystem(5)
    rep = AssignmentReporter(cs)
    assigned = {("A", "S1"): 30, ("B", "S2"): 15}  # A balanced, B +5
    stats = rep.stats(profs, assigned)
    assert stats.expected_total == 40
    assert stats.assigned_total == 45
    assert stats.balanced == 1
    assert stats.overloaded == 1
    assert stats.underloaded == 0
    assert stats.max_overload == 5


def test_reporter_report_lines_contains_summary():
    profs = [Professor(name="A", practice_credits_by_subject={"S1": 6})]
    rep = AssignmentReporter(CreditSystem(5))
    lines = rep.report_lines(profs, {("A", "S1"): 30})
    joined = "\n".join(lines)
    assert "VALIDATION REPORT" in joined
    assert "1 P credit = 5 sessions" in joined


def test_credit_system_quota():
    cs = CreditSystem(5)
    p = Professor(name="A", practice_credits_by_subject={"S1": 3})
    assert cs.quota(p, "S1") == 15
    assert cs.quota(p, "S2") == 0


# --- ExcelGenerator: Teacher view details every session ----------------------
def _build_sessions(prof_name, subject, n_groups=2, n_sessions=3):
    sessions = []
    sid = 0
    for g in range(1, n_groups + 1):
        for k in range(1, n_sessions + 1):
            sessions.append(LabSession(
                id=sid, subject=subject, group_num=g, session_num=k,
                curso_num=2, slot=_slot(), nb_students=12,
                lab_rooms=["Ciencias Exp. I"], window=WeekWindow(4, 14),
                week=3 + k, professor=prof_name,
            ))
            sid += 1
    return sessions


def test_generator_teacher_view_details_each_session(tmp_path):
    prof = Professor(name="Prof X", practice_credits_by_subject={"S1_Test": 2})
    sessions = _build_sessions("Prof X", "S1_Test", n_groups=2, n_sessions=3)
    groups = [
        Group(subject="S1_Test", group_num=g, semester=1, curso_num=2,
            num_sessions=3, slot=_slot(), nb_students=12,
            lab_rooms=["Ciencias Exp. I"], window=WeekWindow(4, 14),
            professor="Prof X")
        for g in (1, 2)
    ]
    gen = ExcelGenerator(CreditSystem(5))
    out = str(tmp_path / "out.xlsx")
    gen.generate(out, groups, sessions, [prof])

    wb = openpyxl.load_workbook(out)
    assert set(["Lab groups", "Subject view", "Teacher view",
                "Validation"]).issubset(set(wb.sheetnames))

    ws = wb["Teacher view"]
    schedule = None
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Prof X" and row[1] == "S1_Test":
            schedule = row[6]
    assert schedule is not None
    # Each session detailed individually (6 lines = 2 groups x 3)
    detail_lines = [ln for ln in schedule.split("\n") if ln.strip()]
    assert len(detail_lines) == 6
    assert "G1 Session 1:" in schedule
    assert "G2 Session 3:" in schedule
    assert "08:30-10:30" in schedule
    assert "Ciencias Exp. I" in schedule


def test_generator_subject_view_single_professor(tmp_path):
    prof = Professor(name="Prof Y", practice_credits_by_subject={"S1_Test": 2})
    sessions = _build_sessions("Prof Y", "S1_Test", n_groups=2, n_sessions=2)
    groups = [
        Group(subject="S1_Test", group_num=g, semester=1, curso_num=2,
            num_sessions=2, slot=_slot(), nb_students=12,
            lab_rooms=["Ciencias Exp. I"], window=WeekWindow(4, 14),
            professor="Prof Y")
        for g in (1, 2)
    ]
    gen = ExcelGenerator(CreditSystem(5))
    out = str(tmp_path / "out.xlsx")
    gen.generate(out, groups, sessions, [prof])

    wb = openpyxl.load_workbook(out)
    ws = wb["Subject view"]
    header = [c.value for c in ws[1]]
    assert "Professor" in header
    prof_col = header.index("Professor")
    # Each session row carries a SINGLE professor, with no "+N" option
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = str(row[prof_col])
        assert "+" not in val
        assert val == "Prof Y"
