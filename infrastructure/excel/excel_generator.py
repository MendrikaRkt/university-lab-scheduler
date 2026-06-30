"""excel_generator.py - Generation of the output Excel files.

Produces the "Distribucion_Practicas" workbooks from a resolved schedule
(LabSession objects with assigned weeks) and the list of professors. It
generates the key sheets:

  - "Lab groups"      : composition of the groups
  - "Subject view"    : one row per session with a SINGLE assigned professor
  - "Teacher view"    : view per professor, DETAIL of each session
  - "Validation"      : expected vs assigned report (overloads/underloads)

Rules applied:
  - 1 P credit = N sessions (N = sessions_per_credit of the CreditSystem).
  - Each session has a SINGLE responsible professor (no more "+6" options).
  - The Teacher view sheet details EACH session individually per group
    (day, time, room, week).

All knowledge of the output format is isolated here.
"""
from __future__ import annotations

from collections import defaultdict
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from application.services.credit_system import AssignmentReporter, CreditSystem
from domain.entities.group import Group, LabSession
from domain.entities.professor import Professor

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
OVERLOAD_FILL = PatternFill("solid", fgColor="FCE4D6")    # light orange
UNDERLOAD_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow
OK_FILL = PatternFill("solid", fgColor="E2EFDA")          # light green


class ExcelGenerator:
    """Builds and writes the practice distribution workbooks."""

    def __init__(self, credit_system: CreditSystem, day_names: Optional[List[str]] = None):
        self.credit_system = credit_system
        # Spanish day names are domain data (they match the source timetable).
        self.day_names = day_names or ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
        self.reporter = AssignmentReporter(credit_system)

    # -- public API ---------------------------------------------------------
    def generate(self, path: str, groups: List[Group],
                sessions: List[LabSession], professors: List[Professor]) -> str:
        wb = Workbook()
        wb.remove(wb.active)
        self._sheet_groups(wb, groups)
        self._sheet_subject_view(wb, sessions)
        self._sheet_teacher_view(wb, sessions, professors)
        self._sheet_validation(wb, sessions, professors)
        wb.save(path)
        return path

    # -- "Lab groups" sheet -------------------------------------------------
    def _sheet_groups(self, wb: Workbook, groups: List[Group]) -> None:
        ws = wb.create_sheet("Lab groups")
        headers = ["Subject", "Group", "Program", "Slot", "Num students",
                "Rooms", "Professor"]
        self._write_header(ws, headers)
        for g in sorted(groups, key=lambda x: (x.subject, x.group_num)):
            ws.append([
                g.subject, f"G{g.group_num}", g.program,
                f"{g.slot.day} {g.slot.block_label}",
                g.nb_students, ", ".join(g.lab_rooms),
                g.professor or "-",
            ])
        self._autosize(ws)

    # -- "Subject view" sheet (one row per session, 1 professor) -----------
    def _sheet_subject_view(self, wb: Workbook, sessions: List[LabSession]) -> None:
        ws = wb.create_sheet("Subject view")
        headers = ["Subject", "Group", "Session", "Professor", "Week",
                "Day", "Time", "Room"]
        self._write_header(ws, headers)
        for s in sorted(sessions, key=lambda x: (x.subject, x.group_num, x.session_num)):
            prof = getattr(s, "professor", None) or "-"
            ws.append([
                s.subject, f"G{s.group_num}", s.session_num, prof,
                s.week if s.week is not None else "-",
                s.slot.day.strip(), s.slot.block_label, ", ".join(s.lab_rooms),
            ])
        self._autosize(ws)

    # -- "Teacher view" sheet (DETAIL of each session per professor) -------
    def _sheet_teacher_view(self, wb: Workbook, sessions: List[LabSession],
                            professors: List[Professor]) -> None:
        ws = wb.create_sheet("Teacher view")
        # Explanatory banner
        ws.append(["Teacher view - detail of the laboratory sessions per professor"])
        ws.append([
            f"Convention: 1 P credit = {self.credit_system.sessions_per_credit} "
            "sessions. Each group is assigned to a SINGLE responsible professor, "
            "in proportion to the official P credits."
        ])
        ws.append([])
        header_row = ws.max_row + 1
        headers = ["Professor", "Subject", "Lab credits (P)",
                f"Expected sessions (cred x{self.credit_system.sessions_per_credit})",
                "Assigned groups", "Planned sessions (professor)",
                "Schedule (detail per session)"]
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

        # Index: P credits per (professor, subject)
        credits_map: Dict[tuple, float] = {}
        for p in professors:
            for subj in p.subjects():
                credits_map[(p.name, subj)] = p.practice_credits(subj)

        # Sessions grouped by (professor, subject) then by group
        by_prof_subject: Dict[tuple, Dict[int, List[LabSession]]] = defaultdict(
            lambda: defaultdict(list))
        for s in sessions:
            prof = getattr(s, "professor", None)
            if not prof:
                continue
            by_prof_subject[(prof, s.subject)][s.group_num].append(s)

        # Union of the keys (professors with credits + professors actually planned)
        keys = set(credits_map) | set(by_prof_subject)
        for (prof, subject) in sorted(keys):
            credits = credits_map.get((prof, subject), 0.0)
            expected = self.credit_system.expected_sessions(credits)
            groups_map = by_prof_subject.get((prof, subject), {})
            assigned_groups = sorted(groups_map.keys())
            planned = sum(len(v) for v in groups_map.values())

            schedule = self._format_schedule_detail(groups_map)
            row = [prof, subject, round(credits, 2), expected,
                ", ".join(f"G{g}" for g in assigned_groups) or "-",
                planned, schedule]
            ws.append(row)
            # formatting: line wrap for the detail + gap highlight
            ws.cell(row=ws.max_row, column=len(headers)).alignment = LEFT_WRAP
            delta = planned - expected
            if delta != 0:
                fill = OVERLOAD_FILL if delta > 0 else UNDERLOAD_FILL
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = fill
        self._autosize(ws, detail_col=len(headers))

    def _format_schedule_detail(self,
                                groups_map: Dict[int, List[LabSession]]) -> str:
        """Details EACH session individually per group.

        Format: "G7 Session 1: Miercoles 08:30-10:30 (Ciencias Exp. I) - Week 5"
        """
        lines: List[str] = []
        for group_num in sorted(groups_map.keys()):
            grp_sessions = sorted(groups_map[group_num], key=lambda x: x.session_num)
            for s in grp_sessions:
                day = s.slot.day.strip()
                room = ", ".join(s.lab_rooms)
                base = (f"G{group_num} Session {s.session_num}: "
                        f"{day} {s.slot.block_label} ({room})")
                if s.week is not None:
                    base += f" - Week {s.week}"
                lines.append(base)
        return "\n".join(lines)

    # -- "Validation" sheet (expected vs assigned report) ------------------
    def _sheet_validation(self, wb: Workbook, sessions: List[LabSession],
                        professors: List[Professor]) -> None:
        ws = wb.create_sheet("Validation")
        assigned: Dict[tuple, int] = defaultdict(int)
        for s in sessions:
            prof = getattr(s, "professor", None)
            if prof:
                assigned[(prof, s.subject)] += 1

        stats = self.reporter.stats(professors, assigned)
        ws.append(["Validation report - Professor assignment"])
        ws.append([f"Rule: 1 P credit = {self.credit_system.sessions_per_credit} sessions"])
        ws.append([f"Professors: {stats.professors}",
                f"(professor, subject) pairs: {stats.pairs}"])
        ws.append([f"Expected sessions: {stats.expected_total}",
                f"Assigned sessions: {stats.assigned_total}"])
        ws.append([f"Balanced: {stats.balanced}",
                f"Overloaded: {stats.overloaded}",
                f"Underloaded: {stats.underloaded}"])
        ws.append([f"Max overload gap: +{stats.max_overload}",
                f"Max underload gap: {stats.max_underload}"])
        ws.append([])

        header_row = ws.max_row + 1
        headers = ["Professor", "Subject", "Lab credits (P)", "Expected",
                "Assigned", "Delta", "Status"]
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

        rows = self.reporter.build_rows(professors, assigned)
        credits_map: Dict[tuple, float] = {}
        for p in professors:
            for subj in p.subjects():
                credits_map[(p.name, subj)] = p.practice_credits(subj)

        for r in rows:
            if r.is_balanced:
                status, fill = "Balanced", OK_FILL
            elif r.is_overload:
                status, fill = "Overload", OVERLOAD_FILL
            else:
                status, fill = "Underload", UNDERLOAD_FILL
            ws.append([
                r.professor, r.subject,
                round(credits_map.get((r.professor, r.subject), 0.0), 2),
                r.expected_sessions, r.planned_sessions, r.delta, status,
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
        self._autosize(ws)

    # -- style helpers ------------------------------------------------------
    def _write_header(self, ws, headers: List[str]) -> None:
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

    @staticmethod
    def _autosize(ws, detail_col: Optional[int] = None) -> None:
        for col in ws.columns:
            idx = col[0].column
            if detail_col is not None and idx == detail_col:
                # detail column: generous fixed width (multi-line content)
                ws.column_dimensions[col[0].column_letter].width = 60
                continue
            width = max((len(str(c.value)) for c in col
                        if c.value is not None and "\n" not in str(c.value)),
                        default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 50)
